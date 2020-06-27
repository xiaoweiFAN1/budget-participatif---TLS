import openpyxl
table = openpyxl.load_workbook(r"C:\Users\Asus\Desktop\Projet tuteuré\regle_greedy.xlsx")# ouvrire file, "r" read
sheet = table.worksheets[0] #position de sheet regle dans excel
nbrow = sheet.max_row  #calculer nombre de lignes
nbcol = sheet.max_column  #calculer nombre de colones

def cal_satisfaction():
    sheet.cell(1, 9, 'Satisfaction')  # inserer titre 'Satisfaction' a cellule(1,9) ou I1   cellule(ligne,colone)
    for i in range(2, nbrow+1):
         sheet.cell(i, 9).value = sheet.cell(i, 6).value  #satisfaction = ceoeur totale, dans colones 6
         table.save('regle_greedy.xlsx')

def regle2_cardinalite_greedy():
    # 1. fonction de transformer les valeurs à une liste
    liste_Satisfaction = []  # 3 listes vides pour memoriser les valeurs
    liste_idee = []
    liste_cout = []
    for i in range(2, nbrow + 1):  # position de ligne
        satisfaction = sheet.cell(i, 9).value  #valeur de satisfaction dans colone 9
        idee = sheet.cell(i, 1).value
        cout = int(sheet.cell(i, 8).value)
        liste_idee.append(idee)   # ajouter les valeurs dans  une liste
        liste_Satisfaction.append(satisfaction)
        liste_cout.append(cout)
        liste_satisfaction_rang = sorted(liste_Satisfaction, reverse=True) # Classez les valeurs par ordre décroissant dans la liste satisfaction

    # trouver la position d'idee et de cout de projet, puis calculer cout totale:
    somme_cout = 0
    Budget = 1000000 # budget limit est 1,000,000
    for m in range(0, len(liste_satisfaction_rang)):
      for n in range(0, len(liste_Satisfaction)):
         if liste_satisfaction_rang[m] == liste_Satisfaction[n]:  # trouver la position correspondante pour liste cout et idee
             position = n
             somme_cout = somme_cout + liste_cout[position]  # calculer la somme de cout par greedy
             if somme_cout <= Budget:
                 Budget_reste = Budget - somme_cout
                 print(liste_idee[position], ", Cout:",liste_cout[position], ", Cout total:", somme_cout)
             else: # somme_cout > budget
                 if Budget_reste > liste_cout[position]:
                    Budget_reste = Budget_reste - liste_cout[position]
                    Budget_totale = Budget- Budget_reste
                    print(liste_idee[position],", Cout:", liste_cout[position],", Cout total:",Budget_totale)



print("Selon Regle 2: cardinalité + greedy，projets choisis sont: ")
# call fonction:
cal_satisfaction()
regle2_cardinalite_greedy()
